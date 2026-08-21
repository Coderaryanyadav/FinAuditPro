"""Unit tests for formula injection escaping across CSV and Excel exports."""

import csv

import openpyxl

from finauditpro.domain.export_sanitizer import escape_formula_injection


def test_escape_formula_injection_prefixes() -> None:
    """Verify formula injection characters (=, +, -, @, \\t, \\r) are escaped with single quote ('')."""
    malicious_inputs = [
        ("=cmd|'/c calc'!A1", "'=cmd|'/c calc'!A1"),
        ("+1+1", "'+1+1"),
        ("-100", "'-100"),
        ("@SUM(A1:A10)", "'@SUM(A1:A10)"),
        ("\tcalc.exe", "'\tcalc.exe"),
        ("\rcalc.exe", "'\rcalc.exe"),
    ]

    for raw, expected in malicious_inputs:
        escaped = escape_formula_injection(raw)
        assert escaped == expected

    # Safe inputs must be untouched
    safe_inputs = ["Normal Text", "Vendor ABC", 123, 45.67, True, None]
    for safe in safe_inputs:
        assert escape_formula_injection(safe) == safe


def test_xlsx_export_contains_escaped_cells(tmp_path) -> None:
    """Verify openpyxl workbook exports malicious cell content safely prefixed with single quote."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    raw_cell = "=SUM(1,2)"
    escaped_cell = escape_formula_injection(raw_cell)
    ws.append(["Title", escaped_cell])

    out_file = tmp_path / "test_injection.xlsx"
    wb.save(str(out_file))

    # Re-open workbook and check cell value
    wb_read = openpyxl.load_workbook(str(out_file))
    sheet_read = wb_read.active
    val = sheet_read.cell(row=1, column=2).value
    assert val == "'=SUM(1,2)"


def test_csv_export_contains_escaped_cells(tmp_path) -> None:
    """Verify csv exports malicious cell content safely prefixed with single quote."""
    out_file = tmp_path / "test_injection.csv"
    raw_cell = "=cmd|'/c calc'!A1"
    escaped_cell = escape_formula_injection(raw_cell)

    with open(out_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Title", escaped_cell])

    with open(out_file, encoding="utf-8") as f:
        content = f.read()

    assert "'=cmd|'/c calc'!A1" in content
