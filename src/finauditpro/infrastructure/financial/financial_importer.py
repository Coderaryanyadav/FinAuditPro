"""Financial data importer for Excel/CSV files with Decimal currency parsing, day-first date parsing, and CSV formula injection protection."""

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from finauditpro.domain.financial_entities import (
    BankTransaction,
    LedgerEntry,
    RowError,
    TrialBalanceLine,
    TrialBalanceSummary,
)
from finauditpro.domain.value_objects import Money
from finauditpro.infrastructure.financial.currency_parser import (
    parse_indian_currency,
    parse_indian_date,
    sanitize_export_cell,
)

__all__ = [
    "FinancialImportError",
    "FinancialImporter",
    "ImportResult",
    "parse_indian_currency",
    "parse_indian_date",
    "sanitize_export_cell",
]


class FinancialImportError(Exception):
    """Raised when dataset file reading or column mapping validation fails."""

    pass


@dataclass
class ImportResult:
    total_rows: int
    valid_rows: list[Any]
    errors: list[RowError]
    summary: TrialBalanceSummary | None = None


class FinancialImporter:
    """Importer converting raw tabular data into validated domain models."""

    @staticmethod
    def read_tabular_rows(file_path: Path) -> tuple[list[str], list[dict[str, str]]]:
        """Read raw Excel or CSV rows as string dictionaries using openpyxl / csv."""
        ext = file_path.suffix.lower()
        rows: list[dict[str, str]] = []
        headers: list[str] = []

        if ext in (".xlsx", ".xls"):
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)

            sheet = wb.active
            iter_rows = sheet.iter_rows(values_only=True)

            header_row = next(iter_rows, None)
            if not header_row:
                raise FinancialImportError("Empty Excel sheet.")

            headers = [str(h or "").strip() for h in header_row]

            for row_tuple in iter_rows:
                if not any(row_tuple):
                    continue
                row_dict = {
                    headers[i]: str(row_tuple[i]).strip()
                    if i < len(row_tuple) and row_tuple[i] is not None
                    else ""
                    for i in range(len(headers))
                }
                rows.append(row_dict)

            wb.close()
        elif ext == ".csv":
            with file_path.open("r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                header_row = next(reader, None)
                if not header_row:
                    raise FinancialImportError("Empty CSV file.")

                headers = [str(h or "").strip() for h in header_row]
                for row_tuple in reader:
                    if not any(row_tuple):
                        continue
                    row_dict = {
                        headers[i]: row_tuple[i].strip() if i < len(row_tuple) else ""
                        for i in range(len(headers))
                    }
                    rows.append(row_dict)
        else:
            raise FinancialImportError(f"Unsupported dataset extension: '{ext}'")

        return headers, rows

    @classmethod
    def import_trial_balance(
        cls, dataset_id: str, rows: list[dict[str, str]], mappings: dict[str, str]
    ) -> ImportResult:
        valid_lines: list[TrialBalanceLine] = []
        errors: list[RowError] = []

        col_code = mappings.get("account_code")
        col_name = mappings.get("account_name", "Account Name")
        col_type = mappings.get("account_type")
        col_op_dr = mappings.get("opening_dr")
        col_op_cr = mappings.get("opening_cr")
        col_dr = mappings.get("debit")
        col_cr = mappings.get("credit")
        col_cl_dr = mappings.get("closing_dr")
        col_cl_cr = mappings.get("closing_cr")

        for idx, r in enumerate(rows, start=2):
            acc_name = r.get(col_name, "").strip() if col_name else ""
            if not acc_name and col_code:
                acc_name = r.get(col_code, "").strip()

            if not acc_name:
                errors.append(
                    RowError(
                        row_no=idx,
                        column_name="account_name",
                        raw_value="",
                        error_reason="Missing Account Name",
                    )
                )
                continue

            try:
                op_dr = parse_indian_currency(r.get(col_op_dr)).paise if col_op_dr else 0
                op_cr = parse_indian_currency(r.get(col_op_cr)).paise if col_op_cr else 0
                dr = parse_indian_currency(r.get(col_dr)).paise if col_dr else 0
                cr = parse_indian_currency(r.get(col_cr)).paise if col_cr else 0
                cl_dr = parse_indian_currency(r.get(col_cl_dr)).paise if col_cl_dr else 0
                cl_cr = parse_indian_currency(r.get(col_cl_cr)).paise if col_cl_cr else 0

                valid_lines.append(
                    TrialBalanceLine(
                        dataset_id=dataset_id,
                        source_row_no=idx,
                        account_code=r.get(col_code, "").strip() if col_code else None,
                        account_name=acc_name,
                        account_type=r.get(col_type, "").strip() if col_type else None,
                        opening_dr_paise=op_dr,
                        opening_cr_paise=op_cr,
                        debit_paise=dr,
                        credit_paise=cr,
                        closing_dr_paise=cl_dr,
                        closing_cr_paise=cl_cr,
                        raw_values=r,
                    )
                )
            except ValueError as val_ex:
                errors.append(
                    RowError(
                        row_no=idx, column_name="amount", raw_value=str(r), error_reason=str(val_ex)
                    )
                )

        total_op_dr = sum(line.opening_dr_paise for line in valid_lines)
        total_op_cr = sum(line.opening_cr_paise for line in valid_lines)
        total_dr = sum(line.debit_paise for line in valid_lines)
        total_cr = sum(line.credit_paise for line in valid_lines)
        total_cl_dr = sum(line.closing_dr_paise for line in valid_lines)
        total_cl_cr = sum(line.closing_cr_paise for line in valid_lines)

        summary = TrialBalanceSummary(
            total_opening_dr_paise=total_op_dr,
            total_opening_cr_paise=total_op_cr,
            total_debit_paise=total_dr,
            total_credit_paise=total_cr,
            total_closing_dr_paise=total_cl_dr,
            total_closing_cr_paise=total_cl_cr,
        )

        if not summary.is_period_balanced and (total_dr > 0 or total_cr > 0):
            diff_paise = summary.period_discrepancy_paise
            diff_money = Money(paise=abs(diff_paise)).format_indian()
            errors.append(
                RowError(
                    row_no=1,
                    column_name="debit/credit",
                    raw_value=f"Debits: {Money(paise=total_dr).format_indian()}, Credits: {Money(paise=total_cr).format_indian()}",
                    error_reason=(
                        f"Trial Balance period movement out of balance by {diff_money} ({diff_paise} paise)."
                    ),
                )
            )

        if not summary.is_closing_balanced and (total_cl_dr > 0 or total_cl_cr > 0):
            diff_paise = summary.closing_discrepancy_paise
            diff_money = Money(paise=abs(diff_paise)).format_indian()
            errors.append(
                RowError(
                    row_no=1,
                    column_name="closing_dr/closing_cr",
                    raw_value=f"Closing Dr: {Money(paise=total_cl_dr).format_indian()}, Closing Cr: {Money(paise=total_cl_cr).format_indian()}",
                    error_reason=(
                        f"Trial Balance closing balances out of balance by {diff_money} ({diff_paise} paise)."
                    ),
                )
            )

        return ImportResult(
            total_rows=len(rows),
            valid_rows=valid_lines,
            errors=errors,
            summary=summary,
        )

    @classmethod
    def import_general_ledger(
        cls, dataset_id: str, rows: list[dict[str, str]], mappings: dict[str, str]
    ) -> ImportResult:
        valid_entries: list[LedgerEntry] = []
        errors: list[RowError] = []

        col_date = mappings.get("date")
        col_type = mappings.get("voucher_type")
        col_no = mappings.get("voucher_number")
        col_code = mappings.get("account_code")
        col_name = mappings.get("account_name")
        col_dr = mappings.get("debit")
        col_cr = mappings.get("credit")
        col_narr = mappings.get("narration")
        col_ref = mappings.get("reference")
        col_by = mappings.get("created_by")

        for idx, r in enumerate(rows, start=2):
            dt_str = None
            if col_date and r.get(col_date):
                try:
                    dt_str = parse_indian_date(r.get(col_date))
                except ValueError as dt_ex:
                    errors.append(
                        RowError(
                            row_no=idx,
                            column_name="date",
                            raw_value=str(r.get(col_date)),
                            error_reason=str(dt_ex),
                        )
                    )
                    continue

            try:
                dr = parse_indian_currency(r.get(col_dr)).paise if col_dr else 0
                cr = parse_indian_currency(r.get(col_cr)).paise if col_cr else 0

                acc_name = r.get(col_name, "").strip() if col_name else ""
                acc_code = r.get(col_code, "").strip() if col_code else None

                valid_entries.append(
                    LedgerEntry(
                        dataset_id=dataset_id,
                        source_row_no=idx,
                        entry_date=dt_str,
                        voucher_type=r.get(col_type, "").strip() if col_type else None,
                        voucher_number=r.get(col_no, "").strip() if col_no else None,
                        account_code=acc_code,
                        account_name=acc_name,
                        debit_paise=dr,
                        credit_paise=cr,
                        narration=r.get(col_narr, "").strip() if col_narr else None,
                        reference=r.get(col_ref, "").strip() if col_ref else None,
                        created_by_raw=r.get(col_by, "").strip() if col_by else None,
                        raw_values=r,
                    )
                )
            except ValueError as val_ex:
                errors.append(
                    RowError(
                        row_no=idx, column_name="amount", raw_value=str(r), error_reason=str(val_ex)
                    )
                )

        return ImportResult(total_rows=len(rows), valid_rows=valid_entries, errors=errors)

    @classmethod
    def import_bank_statement(
        cls, dataset_id: str, rows: list[dict[str, str]], mappings: dict[str, str]
    ) -> ImportResult:
        valid_txns: list[BankTransaction] = []
        errors: list[RowError] = []

        col_date = mappings.get("date")
        col_val_date = mappings.get("value_date")
        col_txn_id = mappings.get("transaction_id")
        col_desc = mappings.get("description", mappings.get("narration"))
        col_dr = mappings.get("debit")
        col_cr = mappings.get("credit")
        col_bal = mappings.get("balance")
        col_ref = mappings.get("reference")

        for idx, r in enumerate(rows, start=2):
            dt_str = None
            if col_date and r.get(col_date):
                try:
                    dt_str = parse_indian_date(r.get(col_date))
                except ValueError as dt_ex:
                    errors.append(
                        RowError(
                            row_no=idx,
                            column_name="date",
                            raw_value=str(r.get(col_date)),
                            error_reason=str(dt_ex),
                        )
                    )
                    continue

            val_dt_str = None
            if col_val_date and r.get(col_val_date):
                try:
                    val_dt_str = parse_indian_date(r.get(col_val_date))
                except Exception:
                    val_dt_str = None

            try:
                dr = parse_indian_currency(r.get(col_dr)).paise if col_dr else 0
                cr = parse_indian_currency(r.get(col_cr)).paise if col_cr else 0
                bal = parse_indian_currency(r.get(col_bal)).paise if col_bal else 0

                desc = r.get(col_desc, "").strip() if col_desc else f"Bank Transaction #{idx}"

                valid_txns.append(
                    BankTransaction(
                        dataset_id=dataset_id,
                        source_row_no=idx,
                        txn_date=dt_str,
                        value_date=val_dt_str,
                        txn_id=r.get(col_txn_id, "").strip() if col_txn_id else None,
                        description=desc,
                        debit_paise=dr,
                        credit_paise=cr,
                        balance_paise=bal,
                        reference=r.get(col_ref, "").strip() if col_ref else None,
                        raw_values=r,
                    )
                )
            except ValueError as val_ex:
                errors.append(
                    RowError(
                        row_no=idx, column_name="amount", raw_value=str(r), error_reason=str(val_ex)
                    )
                )

        return ImportResult(total_rows=len(rows), valid_rows=valid_txns, errors=errors)
