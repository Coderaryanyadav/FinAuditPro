"""Document text and table extractors for born-digital PDFs (pdfplumber), scanned OCR (pypdfium2 + pytesseract), XLSX, CSV, Text, and Images."""

import csv
from dataclasses import dataclass
from pathlib import Path

from finauditpro.domain.document_entities import TextSourceEnum


@dataclass
class ExtractedTable:
    page_number: int
    table_index: int
    rows: list[list[str]]
    bbox: list[float] | None = None


@dataclass
class ExtractedPage:
    page_number: int
    text: str
    text_source: TextSourceEnum = TextSourceEnum.BORN_DIGITAL
    ocr_applied: bool = False
    confidence: float = 1.0
    layout_json: str | None = None


class DocumentExtractorError(Exception):
    """Raised when text or OCR extraction fails."""
    pass


def get_available_tesseract_languages() -> list[str]:
    """Return list of available Tesseract OCR language packs installed on the system."""
    try:
        import pytesseract
        return list(pytesseract.get_languages())
    except Exception:
        return []


def extract_pdf_pages_and_tables(file_path: Path) -> tuple[list[ExtractedPage], list[ExtractedTable]]:
    """Extract pages and tables from PDF. Use pdfplumber for born-digital, pypdfium2 + pytesseract for scanned pages."""
    pages: list[ExtractedPage] = []
    tables: list[ExtractedTable] = []

    # 1. Born-Digital Extraction via pdfplumber
    pdfplumber_available = False
    try:
        import pdfplumber
        pdfplumber_available = True


        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                raw_text = (page.extract_text() or "").strip()

                # Extract Tables
                extracted_tbls = page.extract_tables() or []
                for t_idx, tbl in enumerate(extracted_tbls):
                    clean_rows = [[str(cell or "").strip() for cell in row] for row in tbl]
                    tables.append(ExtractedTable(page_number=i, table_index=t_idx, rows=clean_rows))

                # Check if page has sufficient born-digital text
                if len(raw_text) >= 15:
                    pages.append(
                        ExtractedPage(
                            page_number=i,
                            text=raw_text,
                            text_source=TextSourceEnum.BORN_DIGITAL,
                            ocr_applied=False,
                            confidence=1.0,
                        )
                    )
                else:
                    # Mark page for OCR fallback
                    pages.append(
                        ExtractedPage(
                            page_number=i,
                            text="",
                            text_source=TextSourceEnum.OCR,
                            ocr_applied=False,
                            confidence=0.0,
                        )
                    )
    except Exception:
        pdfplumber_available = False

    # 2. OCR Fallback for empty/scanned pages via pypdfium2 + pytesseract
    needs_ocr = any(p.text_source == TextSourceEnum.OCR or not p.text for p in pages) or not pdfplumber_available

    if needs_ocr:
        try:
            import pypdfium2
            import pytesseract
            from PIL import Image

            pdf_doc = pypdfium2.PdfDocument(file_path)
            num_pages = len(pdf_doc)

            # If pages list was empty, populate empty page entries
            if not pages:
                pages = [
                    ExtractedPage(page_number=i, text="", text_source=TextSourceEnum.OCR, ocr_applied=False, confidence=0.0)
                    for i in range(1, num_pages + 1)
                ]

            avail_langs = get_available_tesseract_languages()
            lang = "eng" if "eng" in avail_langs else None

            for i, pdf_page in enumerate(pdf_doc, start=1):
                page_entry = next((p for p in pages if p.page_number == i), None)
                if page_entry and (page_entry.text_source == TextSourceEnum.OCR or not page_entry.text):
                    # Render page to raster bitmap at 200 DPI
                    pil_img = pdf_page.render(scale=2.0).to_pil()

                    # Run Tesseract OCR with detailed confidence metrics
                    ocr_data = pytesseract.image_to_data(pil_img, lang=lang, output_type=pytesseract.Output.DICT)
                    conf_values = [int(c) for c in ocr_data.get("conf", []) if int(c) >= 0]
                    avg_conf = (sum(conf_values) / (len(conf_values) * 100.0)) if conf_values else 0.50

                    extracted_str = pytesseract.image_to_string(pil_img, lang=lang).strip()

                    page_entry.text = extracted_str
                    page_entry.text_source = TextSourceEnum.OCR
                    page_entry.ocr_applied = True
                    page_entry.confidence = round(avg_conf, 4)

            pdf_doc.close()
        except ImportError as ex:
            # If pypdfium2/pytesseract missing for scanned pages, report honest failure
            for p in pages:
                if not p.text:
                    p.text = f"[Scanned PDF Page {p.page_number}: OCR dependency unavailable ({ex})]"
                    p.confidence = 0.0

    return pages, tables


def extract_image_ocr_pages(file_path: Path) -> list[ExtractedPage]:
    """Extract text and real OCR confidence score from single image files."""
    try:
        import pytesseract
        from PIL import Image


        img = Image.open(file_path)
        avail_langs = get_available_tesseract_languages()
        lang = "eng" if "eng" in avail_langs else None

        ocr_data = pytesseract.image_to_data(img, lang=lang, output_type=pytesseract.Output.DICT)
        conf_values = [int(c) for c in ocr_data.get("conf", []) if int(c) >= 0]
        avg_conf = (sum(conf_values) / (len(conf_values) * 100.0)) if conf_values else 0.50

        text = pytesseract.image_to_string(img, lang=lang).strip()

        return [
            ExtractedPage(
                page_number=1,
                text=text,
                text_source=TextSourceEnum.OCR,
                ocr_applied=True,
                confidence=round(avg_conf, 4),
            )
        ]
    except Exception as ex:
        raise DocumentExtractorError(f"Image OCR processing failed: {ex}") from ex


def extract_csv_pages(file_path: Path) -> list[ExtractedPage]:
    """Extract CSV file contents formatted into a page."""
    rows: list[str] = []
    with file_path.open("r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(" | ".join(row))

    content = "\n".join(rows)
    return [ExtractedPage(page_number=1, text=content, text_source=TextSourceEnum.BORN_DIGITAL, ocr_applied=False, confidence=1.0)]


def extract_excel_pages_and_tables(file_path: Path) -> tuple[list[ExtractedPage], list[ExtractedTable]]:
    """Extract Excel workbook sheets as pages and structured tables using openpyxl."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)

        pages: list[ExtractedPage] = []
        tables: list[ExtractedTable] = []

        for i, sheet_name in enumerate(wb.sheetnames, start=1):
            sheet = wb[sheet_name]
            lines: list[str] = [f"=== Sheet: {sheet_name} ==="]
            sheet_rows: list[list[str]] = []

            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(val).strip() if val is not None else "" for val in row]
                if any(row_vals):
                    lines.append(" | ".join(row_vals))
                    sheet_rows.append(row_vals)

            pages.append(
                ExtractedPage(
                    page_number=i,
                    text="\n".join(lines),
                    text_source=TextSourceEnum.BORN_DIGITAL,
                    ocr_applied=False,
                    confidence=1.0,
                )
            )

            if sheet_rows:
                tables.append(ExtractedTable(page_number=i, table_index=0, rows=sheet_rows))

        wb.close()
        return pages, tables
    except Exception as ex:
        raise DocumentExtractorError(f"Excel extraction error: {ex}") from ex


def extract_text_pages(file_path: Path) -> list[ExtractedPage]:
    """Extract plain text or markdown file content."""
    text = file_path.read_text(encoding="utf-8", errors="replace")
    return [ExtractedPage(page_number=1, text=text.strip(), text_source=TextSourceEnum.BORN_DIGITAL, ocr_applied=False, confidence=1.0)]


def extract_document_content(file_path: Path, mime_type: str) -> tuple[list[ExtractedPage], list[ExtractedTable]]:
    """Route document to appropriate extractor based on mime type or extension."""
    ext = file_path.suffix.lower()

    if mime_type == "application/pdf" or ext == ".pdf":
        return extract_pdf_pages_and_tables(file_path)
    if mime_type == "text/csv" or ext == ".csv":
        return extract_csv_pages(file_path), []
    if ext in (".xlsx", ".xls"):
        return extract_excel_pages_and_tables(file_path)
    if mime_type.startswith("image/") or ext in (".png", ".jpg", ".jpeg", ".tiff"):
        return extract_image_ocr_pages(file_path), []
    if mime_type.startswith("text/") or ext in (".txt", ".md", ".log", ".json"):
        return extract_text_pages(file_path), []

    # Fallback to plain text
    try:
        return extract_text_pages(file_path), []
    except Exception:
        return (
            [ExtractedPage(page_number=1, text=f"[Binary Document Content: '{file_path.name}']", text_source=TextSourceEnum.BORN_DIGITAL, ocr_applied=False, confidence=1.0)],
            [],
        )
