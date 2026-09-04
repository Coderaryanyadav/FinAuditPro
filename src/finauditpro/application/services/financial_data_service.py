"""Financial data import and normalization service."""

import csv
from pathlib import Path
from uuid import uuid4

from finauditpro.application.financial_dtos import ImportDatasetDTO, InspectFileResultDTO
from finauditpro.application.security.engagement_lock_guard import assert_engagement_not_locked
from finauditpro.domain.entities import AuditEvent
from finauditpro.domain.exceptions import EntityNotFoundError
from finauditpro.domain.financial_entities import FinancialDataset, FinancialRecord
from finauditpro.infrastructure.analytics.column_detector import detect_column_mappings
from finauditpro.infrastructure.documents.document_security import calculate_sha256
from finauditpro.infrastructure.persistence.database import DatabaseManager
from finauditpro.infrastructure.persistence.repositories import (
    AuditEventRepository,
    EngagementRepository,
    FinancialDataRepository,
)


class FinancialDataService:
    """Service handling financial dataset file inspection, column mapping, and record normalization."""

    def __init__(self, db_manager: DatabaseManager) -> None:
        self.db_manager = db_manager

    def inspect_file(self, file_path: str | Path) -> InspectFileResultDTO:
        """Read CSV or Excel headers and sample preview rows to auto-detect column mappings."""
        path = Path(file_path)
        if not path.is_file():
            raise EntityNotFoundError("Financial Data File", str(file_path))

        headers: list[str] = []
        preview_rows: list[dict[str, str]] = []

        ext = path.suffix.lower()
        if ext == ".csv":
            with path.open("r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                try:
                    headers = next(reader)
                except StopIteration:
                    headers = []

                for _ in range(10):
                    try:
                        row = next(reader)
                        preview_rows.append(
                            {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                        )
                    except StopIteration:
                        break
        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl

                wb = openpyxl.load_workbook(path, data_only=True)

                sheet = wb.active
                iter_rows = sheet.iter_rows(values_only=True)
                try:
                    raw_headers = next(iter_rows)
                    headers = [
                        str(h) if h is not None else f"Column_{i + 1}"
                        for i, h in enumerate(raw_headers)
                    ]
                    for _ in range(10):
                        row = next(iter_rows)
                        row_dict = {
                            headers[i]: str(row[i]) if i < len(row) and row[i] is not None else ""
                            for i in range(len(headers))
                        }
                        preview_rows.append(row_dict)
                except StopIteration:
                    headers = []
                wb.close()
            except Exception:
                headers = ["Column_1"]

        mappings = detect_column_mappings(headers, preview_rows)
        return InspectFileResultDTO(
            file_path=str(path.resolve()),
            headers=headers,
            suggested_mappings=mappings,
            preview_rows=preview_rows,
        )

    def import_financial_dataset(self, dto: ImportDatasetDTO) -> FinancialDataset:
        """Import financial data file, apply column mappings, create normalized records, and persist."""
        path = Path(dto.file_path)
        if not path.is_file():
            raise EntityNotFoundError("Financial Data File", dto.file_path)

        with self.db_manager.session_scope() as session:
            eng_repo = EngagementRepository(session)
            eng = eng_repo.get_by_id(dto.engagement_id)
            if not eng:
                raise EntityNotFoundError("Engagement", dto.engagement_id)
            assert_engagement_not_locked(eng)

        content_hash = calculate_sha256(path)

        # Inspect and parse file rows
        inspection = self.inspect_file(path)
        mappings = dto.column_mappings or inspection.suggested_mappings

        records: list[FinancialRecord] = []
        raw_rows: list[dict[str, str]] = []

        ext = path.suffix.lower()
        if ext == ".csv":
            with path.open("r", encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    raw_rows.append(r)
        elif ext in (".xlsx", ".xls"):
            import openpyxl

            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.active
            iter_rows = sheet.iter_rows(values_only=True)
            try:
                raw_h = [
                    str(h) if h is not None else f"Column_{i + 1}"
                    for i, h in enumerate(next(iter_rows))
                ]
                for r in iter_rows:
                    row_dict = {
                        raw_h[i]: str(r[i]) if i < len(r) and r[i] is not None else ""
                        for i in range(len(raw_h))
                    }
                    raw_rows.append(row_dict)
            except StopIteration:
                pass
            wb.close()

        # Normalize rows to FinancialRecord entities
        m_date = mappings.get("date")
        m_code = mappings.get("account_code")
        m_name = mappings.get("account_name")
        m_debit = mappings.get("debit")
        m_credit = mappings.get("credit")
        m_amount = mappings.get("amount")
        m_narration = mappings.get("narration")
        m_counterparty = mappings.get("counterparty_name")
        m_gstin = mappings.get("counterparty_gstin")
        m_inv = mappings.get("invoice_number")

        dataset_id = str(uuid4())

        for idx, row in enumerate(raw_rows, start=1):

            def _get_float(key: str | None, target_row: dict[str, str]) -> float:
                if not key or key not in target_row:
                    return 0.0
                val_str = str(target_row[key]).replace(",", "").strip()
                try:
                    return float(val_str)
                except ValueError:
                    return 0.0

            debit_val = _get_float(m_debit, row)
            credit_val = _get_float(m_credit, row)
            amount_val = _get_float(m_amount, row)
            if amount_val == 0.0 and (debit_val > 0 or credit_val > 0):
                amount_val = debit_val - credit_val

            rec = FinancialRecord(
                dataset_id=dataset_id,
                row_index=idx,
                transaction_id=str(row.get(m_inv or "", f"ROW-{idx}")),
                date=str(row.get(m_date or "", "") or None),
                account_code=str(row.get(m_code or "", "") or None),
                account_name=str(row.get(m_name or "", "") or None),
                debit=debit_val,
                credit=credit_val,
                amount=amount_val,
                narration=str(row.get(m_narration or "", "") or None),
                counterparty_name=str(row.get(m_counterparty or "", "") or None),
                counterparty_gstin=str(row.get(m_gstin or "", "") or None),
                invoice_number=str(row.get(m_inv or "", "") or None),
            )
            records.append(rec)

        dataset = FinancialDataset(
            id=dataset_id,
            engagement_id=dto.engagement_id,
            dataset_name=dto.dataset_name,
            dataset_type=dto.dataset_type,
            file_path=str(path.resolve()),
            content_hash=content_hash,
            column_mappings=mappings,
            row_count=len(records),
        )

        with self.db_manager.session_scope() as session:
            fin_repo = FinancialDataRepository(session)
            created_ds = fin_repo.add_dataset(dataset)
            fin_repo.add_records(records)

            audit_repo = AuditEventRepository(session)
            audit_repo.add(
                AuditEvent(
                    engagement_id=dto.engagement_id,
                    actor="System",
                    action="Financial Dataset Imported",
                    details=f"Imported '{dataset.dataset_name}' ({len(records)} rows, Type: {dataset.dataset_type.value})",
                )
            )

        return created_ds

    def list_datasets_for_engagement(self, engagement_id: str) -> list[FinancialDataset]:
        with self.db_manager.session_scope() as session:
            repo = FinancialDataRepository(session)
            return repo.list_datasets_by_engagement(engagement_id)

    def get_dataset_records(self, dataset_id: str) -> list[FinancialRecord]:
        with self.db_manager.session_scope() as session:
            repo = FinancialDataRepository(session)
            return repo.get_dataset_records(dataset_id)
