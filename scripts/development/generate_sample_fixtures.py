"""
Generate statutory audit sample test files (PDF, XLSX, CSV, JSON)
for end-to-end testing and validation of FinAuditPro.
"""

import csv
import json
from pathlib import Path

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def generate_all_test_fixtures(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Creating sample statutory test files in: {output_dir}")

    # 1. Trial Balance XLSX
    wb_tb = openpyxl.Workbook()
    ws_tb = wb_tb.active
    ws_tb.title = "Trial_Balance_FY24_25"
    ws_tb.append(["Account Code", "Account Name", "Schedule III Classification", "Debit (INR)", "Credit (INR)"])
    tb_data = [
        ["1001", "Equity Share Capital", "Shareholders' Funds", 0.0, 10000000.0],
        ["1002", "Retained Earnings / Surplus", "Reserves & Surplus", 0.0, 15000000.0],
        ["2001", "HDFC Bank Term Loan", "Long-Term Borrowings", 0.0, 25000000.0],
        ["2002", "Sundry Creditors (MSME)", "Trade Payables", 0.0, 4500000.0],
        ["2003", "Sundry Creditors (Non-MSME)", "Trade Payables", 0.0, 10500000.0],
        ["3001", "Plant & Machinery (Gross)", "Property, Plant & Equipment", 35000000.0, 0.0],
        ["3002", "Accumulated Depreciation", "Property, Plant & Equipment", 0.0, 7000000.0],
        ["3003", "Raw Material Inventory", "Inventories", 12000000.0, 0.0],
        ["3004", "Sundry Debtors", "Trade Receivables", 18000000.0, 0.0],
        ["3005", "HDFC Current Account", "Cash & Cash Equivalents", 7000000.0, 0.0],
        ["4001", "Revenue from Operations", "Revenue", 0.0, 80000000.0],
        ["5001", "Raw Material Purchases", "Expenses", 45000000.0, 0.0],
        ["5002", "Employee Salaries & Bonus", "Employee Benefit Expenses", 15000000.0, 0.0],
        ["5003", "Bank Interest & Charges", "Finance Costs", 2500000.0, 0.0],
        ["5004", "Depreciation Expense", "Depreciation", 3500000.0, 0.0],
        ["5005", "Power & Operational Costs", "Other Expenses", 4000000.0, 0.0],
    ]
    for row in tb_data:
        ws_tb.append(row)
    wb_tb.save(output_dir / "sample_trial_balance_fy24_25.xlsx")
    print("  ✓ sample_trial_balance_fy24_25.xlsx")

    # 2. General Ledger / Voucher Transactions CSV
    with open(output_dir / "sample_general_ledger_vouchers.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Voucher Number", "Date", "Account Name", "Amount (INR)", "Transaction Type", "Narration", "User ID"])
        vouchers = [
            ["V-1001", "2025-04-10", "Raw Material Purchases", 450000.0, "Debit", "Purchase from ABC Tech MSME", "acc_user1"],
            ["V-1002", "2025-05-15", "Employee Salaries", 1250000.0, "Debit", "Monthly salary disbursement", "hr_admin"],
            ["V-1003", "2025-08-20", "Plant & Machinery", 5000000.0, "Debit", "Capitalization of CNC Machine", "partner_ca"],
            ["V-1004", "2025-11-25", "Raw Material Purchases", 450000.0, "Debit", "Duplicate payment test voucher", "acc_user2"],
            ["V-1005", "2026-03-28", "Revenue from Operations", 1000000.0, "Credit", "Invoice S-101 sales to Alpha Corp", "sales_team"],
            ["V-1006", "2026-03-31", "Sundry Debtors", 1800000.0, "Debit", "Year-end debtor closing balance", "acc_user1"],
        ]
        for v in vouchers:
            writer.writerow(v)
    print("  ✓ sample_general_ledger_vouchers.csv")

    # 3. GSTR-2B Inward Supplies CSV
    with open(output_dir / "sample_gstr2b_inward_supplies.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Invoice Number", "Invoice Date", "Supplier GSTIN", "Supplier Name", "Taxable Value (INR)", "Tax Amount (INR)", "ITC Eligibility"])
        gstr_rows = [
            ["INV-001", "2025-04-10", "27AAACB1234F1Z5", "ABC Tech MSME", 100000.0, 18000.0, "Eligible"],
            ["INV-002", "2025-06-12", "27AAACB1234F1Z5", "ABC Tech MSME", 50000.0, 9000.0, "Ineligible Sec 17(5)"],
            ["INV-003", "2025-09-18", "27XYZAB9999F1Z1", "Zenith Logistics Ltd", 200000.0, 36000.0, "Eligible"],
        ]
        for r in gstr_rows:
            writer.writerow(r)
    print("  ✓ sample_gstr2b_inward_supplies.csv")

    # 4. Bank Reconciliation Statement (BRS) XLSX
    wb_brs = openpyxl.Workbook()
    ws_brs = wb_brs.active
    ws_brs.title = "BRS_March_2026"
    ws_brs.append(["Bank Account Number", "Item Type", "Reference / Cheque No", "Entry Date", "Amount (INR)", "Clearance Date", "Audit Status"])
    brs_rows = [
        ["HDFC-0011223344", "Cheque Issued but Not Presented", "CHQ-889901", "2026-03-25", 250000.0, "2026-04-04", "Normal Timing Difference"],
        ["HDFC-0011223344", "Cheque Issued but Not Presented", "CHQ-776655", "2025-11-15", 800000.0, "", "Stale Cheque (>90 Days) - Reversal Required"],
        ["HDFC-0011223344", "Cheque Deposited but Not Credited", "DEP-112233", "2026-03-01", 1500000.0, "", "Delayed Banking (>15 Days) - Inquiry Required"],
    ]
    for r in brs_rows:
        ws_brs.append(r)
    wb_brs.save(output_dir / "sample_bank_reconciliation_statement.xlsx")
    print("  ✓ sample_bank_reconciliation_statement.xlsx")

    # 5. Fixed Asset Register XLSX
    wb_fa = openpyxl.Workbook()
    ws_fa = wb_fa.active
    ws_fa.title = "Fixed_Asset_Register"
    ws_fa.append(["Asset Tag", "Asset Name", "Gross Block (INR)", "Accumulated Dep (INR)", "Net Book Value (INR)", "Location", "Physically Verified", "Title Deeds in Name"])
    fa_rows = [
        ["FA-001", "CNC Lathe Machine #1", 10000000.0, 2000000.0, 8000000.0, "Factory Unit A", "Yes", "Yes"],
        ["FA-002", "Factory Land & Building", 25000000.0, 0.0, 25000000.0, "Pune Plant", "Yes", "Yes"],
        ["FA-003", "Old Diesel Generator", 5000000.0, 6000000.0, -1000000.0, "Factory Unit B", "Yes", "Yes"],
        ["FA-004", "Executive Laptop Series 4", 800000.0, 200000.0, 600000.0, "Head Office", "No", "Yes"],
    ]
    for r in fa_rows:
        ws_fa.append(r)
    wb_fa.save(output_dir / "sample_fixed_asset_register.xlsx")
    print("  ✓ sample_fixed_asset_register.xlsx")

    # 6. Statutory Audit Engagement PDF Document (Mock Signed Letter / Evidence)
    pdf_path = output_dir / "sample_statutory_engagement_letter.pdf"
    doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>STATUTORY AUDIT ENGAGEMENT LETTER (SA 210)</b>", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<b>To the Board of Directors:</b> Reliance Enterprises Private Limited", styles["Normal"]))
    elements.append(Paragraph("<b>Audit Period:</b> Financial Year 2024-25", styles["Normal"]))
    elements.append(Paragraph("<b>Auditor:</b> Apex Statutory Auditors & Co. (FRN: FRN-123456W)", styles["Normal"]))
    elements.append(Spacer(1, 14))

    body_text = (
        "We are pleased to confirm our acceptance and our understanding of this statutory audit engagement "
        "by means of this letter. Our audit will be conducted in accordance with the Standards on Auditing (SAs) "
        "issued by the Institute of Chartered Accountants of India (ICAI) and the provisions of Section 139 & 143 "
        "of the Companies Act, 2013."
    )
    elements.append(Paragraph(body_text, styles["BodyText"]))
    elements.append(Spacer(1, 14))

    scope_data = [
        ["Scope Item", "Statutory Reference", "Responsibility"],
        ["Financial Statement Truth & Fairness", "Section 143(2) CA 2013", "Statutory Auditor"],
        ["Internal Financial Controls (IFCoFR)", "Section 143(3)(i) CA 2013", "Statutory Auditor"],
        ["CARO 2020 Clause Reporting", "Order under Sec 143(11)", "Statutory Auditor"],
        ["Preparation of Financial Books", "Section 134(5) CA 2013", "Management / Board"],
    ]
    t = Table(scope_data, colWidths=[200, 150, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<b>Agreed Audit Fee:</b> ₹5,00,000/- (Rupees Five Lakhs Only) plus applicable GST.", styles["Normal"]))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<i>Signed on behalf of Management and Statutory Auditors.</i>", styles["Italic"]))

    doc.build(elements)
    print("  ✓ sample_statutory_engagement_letter.pdf")

    # 7. Metadata Manifest JSON
    manifest = {
        "manifest_version": "2.0.0",
        "description": "Statutory audit sample test fixtures for FinAuditPro enterprise test suite.",
        "files": [
            {"filename": "sample_trial_balance_fy24_25.xlsx", "format": "XLSX", "purpose": "Trial Balance ingestion & Schedule III classification"},
            {"filename": "sample_general_ledger_vouchers.csv", "format": "CSV", "purpose": "Benford analysis, duplicate payment, and cut-off testing"},
            {"filename": "sample_gstr2b_inward_supplies.csv", "format": "CSV", "purpose": "GSTR-2B vs Purchase register 3-way reconciliation"},
            {"filename": "sample_bank_reconciliation_statement.xlsx", "format": "XLSX", "purpose": "BRS stale cheque (>90d) and delayed banking analytics"},
            {"filename": "sample_fixed_asset_register.xlsx", "format": "XLSX", "purpose": "CARO 2020 Clause 3(i) and negative NBV anomaly testing"},
            {"filename": "sample_statutory_engagement_letter.pdf", "format": "PDF", "purpose": "FTS5 OCR indexing, Evidence DAG, and Bounding Box viewer testing"},
        ]
    }
    with open(output_dir / "test_fixtures_manifest.json", "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)
    print("  ✓ test_fixtures_manifest.json")


if __name__ == "__main__":
    generate_all_test_fixtures(Path("sample"))

